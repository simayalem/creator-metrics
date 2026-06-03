from flask import Flask, render_template, request, send_file, jsonify
import asyncio
import requests
import time
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from playwright.async_api import async_playwright
import os

app = Flask(__name__)

APIFY_TOKEN = "apify_api_QoW4djk70i7ucFtZ0LxJe8uRJyHQPE0bH0mA"

def detect_platform(url):
    if "instagram.com" in url:
        return "Instagram"
    elif "tiktok.com" in url or "vt.tiktok.com" in url:
        return "TikTok"
    elif "youtube.com" in url or "youtu.be" in url:
        return "YouTube"
    elif "x.com" in url or "twitter.com" in url:
        return "X (Twitter)"
    return "Bilinmiyor"

# === INSTAGRAM (Apify) ===
def get_instagram_metrics(url):
    headers = {"Authorization": f"Bearer {APIFY_TOKEN}", "Content-Type": "application/json"}
    run = requests.post(
        "https://api.apify.com/v2/acts/apify~instagram-scraper/runs",
        headers=headers,
        json={"directUrls": [url], "resultsType": "details", "resultsLimit": 1}
    )
    if run.status_code not in [200, 201]:
        return {}
    run_id = run.json()["data"]["id"]
    for _ in range(30):
        time.sleep(5)
        status = requests.get(f"https://api.apify.com/v2/actor-runs/{run_id}", headers=headers).json()["data"]
        if status["status"] == "SUCCEEDED":
            dataset_id = status["defaultDatasetId"]
            results = requests.get(f"https://api.apify.com/v2/datasets/{dataset_id}/items", headers=headers).json()
            if results:
                item = results[0]
                owner = item.get("ownerUsername") or item.get("owner", {}).get("username") or ""
                return {
                    "profile":  owner,
                    "views":    item.get("videoPlayCount") or item.get("playCount") or item.get("viewsCount") or 0,
                    "likes":    item.get("likesCount") or 0,
                    "comments": item.get("commentsCount") or 0,
                    "shares":   item.get("sharesCount") or 0,
                }
            break
        elif status["status"] in ["FAILED", "ABORTED"]:
            break
    return {}

# === X (Twitter) ===
async def get_twitter_metrics(page, url):
    await page.goto(url, wait_until="domcontentloaded", timeout=20000)
    await page.wait_for_timeout(5000)

    profile_match = re.search(r'(?:x|twitter)\.com/([^/]+)/status', url)
    profile = profile_match.group(1) if profile_match else ""
    metrics = {"profile": profile, "views": 0, "likes": 0, "comments": 0, "shares": 0}

    elements = await page.query_selector_all('[aria-label]')
    for el in elements:
        label = await el.get_attribute('aria-label') or ''
        label_lower = label.lower()
        if 'views' in label_lower and 'replies' in label_lower:
            # "33 replies, 27 reposts, 792 likes, 114933 views"
            views_match   = re.search(r'([\d,]+)\s*views', label, re.IGNORECASE)
            replies_match = re.search(r'([\d,]+)\s*repl', label, re.IGNORECASE)
            repost_match  = re.search(r'([\d,]+)\s*repost', label, re.IGNORECASE)
            likes_match   = re.search(r'([\d,]+)\s*like', label, re.IGNORECASE)
            if views_match:   metrics["views"]    = int(views_match.group(1).replace(',', ''))
            if replies_match: metrics["comments"] = int(replies_match.group(1).replace(',', ''))
            if repost_match:  metrics["shares"]   = int(repost_match.group(1).replace(',', ''))
            if likes_match:   metrics["likes"]    = int(likes_match.group(1).replace(',', ''))
            break

    return metrics

# === YOUTUBE ===
async def get_youtube_metrics(page, url):
    await page.goto(url, wait_until="domcontentloaded", timeout=20000)
    await page.wait_for_timeout(6000)
    content = await page.content()
    metrics = {"profile": "", "views": 0, "likes": 0, "comments": 0, "shares": 0}
    view_match = re.search(r'"viewCount":"(\d+)"', content)
    if view_match:
        metrics["views"] = int(view_match.group(1))
    like_match = re.search(r'"likeCount":"(\d+)"', content)
    if like_match:
        metrics["likes"] = int(like_match.group(1))
    # YouTube yorumları lazy load — çekilemiyor
    channel_match = re.search(r'"channelName":"([^"]+)"', content)
    if not channel_match:
        channel_match = re.search(r'"ownerChannelName":"([^"]+)"', content)
    if channel_match:
        metrics["profile"] = channel_match.group(1)
    return metrics

# === TIKTOK ===
async def get_tiktok_metrics(page, url):
    await page.goto(url, wait_until="domcontentloaded", timeout=30000)
    await page.wait_for_timeout(8000)  # TikTok yavaş yükleniyor
    content = await page.content()
    metrics = {"profile": "", "views": 0, "likes": 0, "comments": 0, "shares": 0}
    for key, patterns in {
        "views":    [r'"playCount":(\d+)', r'"play_count":(\d+)'],
        "likes":    [r'"diggCount":(\d+)', r'"like_count":(\d+)'],
        "comments": [r'"commentCount":(\d+)', r'"comment_count":(\d+)'],
        "shares":   [r'"shareCount":(\d+)', r'"share_count":(\d+)'],
    }.items():
        for pattern in patterns:
            match = re.search(pattern, content)
            if match:
                metrics[key] = int(match.group(1))
                break
    # TikTok profil — URL'den
    profile_match = re.search(r'tiktok\.com/@([^/?]+)', page.url)
    if profile_match:
        metrics["profile"] = "@" + profile_match.group(1)
    return metrics

async def scrape_x_youtube(links):
    ua = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    results = {}
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(user_agent=ua)
        page = await ctx.new_page()
        for url in links:
            platform = detect_platform(url)
            try:
                if "Twitter" in platform or "X (" in platform:
                    results[url] = await get_twitter_metrics(page, url)
                elif platform == "YouTube":
                    results[url] = await get_youtube_metrics(page, url)
            except:
                results[url] = {"profile": "", "views": 0, "likes": 0, "comments": 0, "shares": 0}
        await browser.close()
    return results

async def scrape_tiktok(links):
    ua = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    results = {}
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(user_agent=ua)
        page = await ctx.new_page()
        for url in links:
            try:
                results[url] = await get_tiktok_metrics(page, url)
            except Exception as e:
                print(f"TikTok hata: {e}")
                results[url] = {"profile": "", "views": 0, "likes": 0, "comments": 0, "shares": 0}
        await browser.close()
    return results

def process_links(links):
    from concurrent.futures import ThreadPoolExecutor

    instagram_links = [l for l in links if detect_platform(l) == "Instagram"]
    xy_links = [l for l in links if detect_platform(l) in ["X (Twitter)", "YouTube"]]
    tt_links = [l for l in links if detect_platform(l) == "TikTok"]

    results_map = {}

    def run_instagram(ls):
        return {url: get_instagram_metrics(url) for url in ls}

    def run_xy(ls):
        return asyncio.run(scrape_x_youtube(ls))

    def run_tiktok(ls):
        try:
            return asyncio.run(scrape_tiktok(ls))
        except Exception as e:
            print(f"run_tiktok hata: {e}")
            return {url: {"profile": "", "views": 0, "likes": 0, "comments": 0, "shares": 0} for url in ls}

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = []
        if instagram_links:
            futures.append(executor.submit(run_instagram, instagram_links))
        if xy_links:
            futures.append(executor.submit(run_xy, xy_links))
        if tt_links:
            futures.append(executor.submit(run_tiktok, tt_links))
        for f in futures:
            results_map.update(f.result())

    results = []
    for url in links:
        platform = detect_platform(url)
        m = results_map.get(url, {})
        results.append({"link": url, "platform": platform, **m})

    return results

def save_excel(results):
    filename = f"/Users/simay/rapor_app/rapor_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Görüntülenme Raporu"

    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    headers = ["#", "Platform", "Profil", "Görüntülenme", "Beğeni", "Yorum", "Paylaşım", "Link", "Tarih"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    platform_colors = {"Instagram": "E1306C", "X (Twitter)": "1DA1F2", "YouTube": "FF0000", "TikTok": "888888"}
    for i, r in enumerate(results, 1):
        ws.append([
            i,
            r.get("platform", ""),
            r.get("profile", ""),
            r.get("views", 0),
            r.get("likes", 0),
            r.get("comments", 0),
            r.get("shares", 0),
            r.get("link", ""),
            datetime.now().strftime("%d.%m.%Y"),
        ])
        color = platform_colors.get(r.get("platform", ""), "666666")
        for col in range(2, 7):
            ws.cell(row=i+1, column=col).number_format = '#,##0'
        ws.cell(row=i+1, column=2).font = Font(color=color, bold=True)

    for col, width in zip('ABCDEFGHIJ', [5, 15, 20, 16, 12, 12, 12, 60, 12, 12]):
        ws.column_dimensions[col].width = width

    wb.save(filename)
    return filename

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/process", methods=["POST"])
def process():
    data = request.json
    links = [l.strip() for l in data.get("links", []) if l.strip().startswith("http")]
    if not links:
        return jsonify({"error": "Link girilmedi"}), 400
    results = process_links(links)
    filename = save_excel(results)
    return jsonify({
        "results": results,
        "total_views": sum(r.get("views", 0) for r in results),
        "filename": os.path.basename(filename)
    })

@app.route("/download/<filename>")
def download(filename):
    path = f"/Users/simay/rapor_app/{filename}"
    return send_file(path, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=False, port=5001)
